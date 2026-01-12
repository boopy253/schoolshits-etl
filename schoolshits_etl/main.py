import argparse
import re
import time
from copy import copy
from typing import Any, cast

import pandas as pd
from loguru import logger
from openpyxl import load_workbook
from openpyxl.cell import Cell
from openpyxl.worksheet.worksheet import Worksheet


def parse_args() -> argparse.Namespace:
    """Parse command line arguments."""
    parser = argparse.ArgumentParser(description="Populate target Excel template from source data.")

    parser.add_argument("--source", default="source.xlsx", help="Path to input Excel file")
    parser.add_argument("--template", default="target.xlsx", help="Path to template Excel file")
    parser.add_argument("--output", default="target_filled.xlsx", help="Path to output Excel file")
    parser.add_argument("--side-output", default="version_grade_book.xlsx", help="Path to side output Excel file")
    parser.add_argument("--school", default="示例小学")
    parser.add_argument("--year", default="2025")
    parser.add_argument("--start-row", type=int, default=4, help="Row number to start writing data in template")

    return parser.parse_args()


def normalize_text(text: Any) -> str:
    """Normalize text by converting brackets and stripping whitespace."""
    if not isinstance(text, str):
        return ""
    return text.replace("（", "(").replace("）", ")").strip()


def copy_row_style(ws: Worksheet, src_row: int, tgt_row: int, max_col: int) -> None:
    """
    Copy cell styles (font, border, alignment) from source row to target row.
    Essential for maintaining template aesthetics when data exceeds template rows.
    """
    for col in range(1, max_col + 1):
        src = ws.cell(row=src_row, column=col)
        tgt = ws.cell(row=tgt_row, column=col)

        if isinstance(tgt, Cell):
            tgt.value = src.value  # Copy value to ensure no blank cells if logic requires
            if src.has_style:
                tgt._style = copy(src._style)


def load_and_normalize_source(path: str) -> pd.DataFrame:
    """
    Load source Excel and normalize columns based on detected header format.
    Strategies:
    1. '发货单明细': Split '发货数' based on '是否免费'.
    2. '教辅目录': Dynamic column location for student/teacher counts.
    3. '免费教材': Fuzzy match for Name/Version columns.
    4. Default: Standard format.
    """
    logger.info(f"Loading source file: {path}")
    try:
        raw = pd.read_excel(path, header=None)
    except Exception as e:
        logger.error(f"Failed to read source file: {e}")
        raise

    first_cell = str(raw.iloc[0, 0])
    df2: pd.DataFrame | None = None

    if "发货单明细" in first_cell:
        logger.info("Format detected: [Shipment Details] (Header at row 2)")
        df = pd.read_excel(path, header=1)

        def split_row(r: pd.Series) -> pd.Series:
            free = str(r.get("是否免费")).strip() == "是"
            return pd.Series(
                {
                    "书名": r.get("产品名称"),
                    "版别": None,
                    "单价": r.get("定价"),
                    "非免费订数": 0 if free else r.get("发货数"),
                    "免费订数": r.get("发货数") if free else 0,
                }
            )

        df2 = cast(pd.DataFrame, df.apply(split_row, axis=1))

    elif "教辅目录" in first_cell:
        logger.info("Format detected: [Supplement Catalog] (Header at row 2)")
        df = pd.read_excel(path, header=1)
        student_col = df.columns[-2]
        teacher_col = df.columns[-1]

        df2 = pd.DataFrame(
            {
                "书名": df["名称"],
                "版别": df["版本"],
                "单价": df["单价"],
                "非免费订数": df[student_col],
                "免费订数": df[teacher_col],
            }
        )

    elif "免费教材" in first_cell:
        logger.info("Format detected: [Free Textbook Catalog] (Header at row 3)")
        df = pd.read_excel(path, header=2)

        def find_col_by_keywords(df_in: pd.DataFrame, keywords: list[str]) -> Any | None:
            for col in df_in.columns:
                col_str = str(col).replace(" ", "").replace("\u3000", "")
                if all(k in col_str for k in keywords):
                    return col
            return None

        name_col = find_col_by_keywords(df, ["书", "名"])
        version_col = find_col_by_keywords(df, ["版"])

        df2 = pd.DataFrame(
            {
                "书名": df[name_col],
                "版别": df[version_col] if version_col else None,
                "单价": 0,
                "非免费订数": 0,
                "免费订数": df.iloc[:, -1],
            }
        )

    else:
        logger.info("Format detected: [Standard Statistics Table]")
        df = pd.read_excel(path)
        df2 = cast(pd.DataFrame, df[["书名", "版别", "单价", "非免费订数", "免费订数"]].copy())

    # Ensure numeric types
    for col in ["非免费订数", "免费订数"]:
        df2[col] = cast(pd.Series, pd.to_numeric(df2[col], errors="coerce")).fillna(0).astype(int)

    df2["单价"] = cast(pd.Series, pd.to_numeric(df2["单价"], errors="coerce")).fillna(0)

    logger.success(f"Source loaded successfully. Total rows: {len(df2)}")
    return df2


def parse_grade_and_term(book_name: Any) -> tuple[str | None, str | None]:
    """Extract grade (e.g., '一年级') and term (e.g., '上学期') from book name."""
    if not isinstance(book_name, str):
        return None, None

    text = normalize_text(book_name)

    grade_map = {
        "一": "一年级",
        "二": "二年级",
        "三": "三年级",
        "四": "四年级",
        "五": "五年级",
        "六": "六年级",
        "1": "一年级",
        "2": "二年级",
        "3": "三年级",
        "4": "四年级",
        "5": "五年级",
        "6": "六年级",
    }

    grade: str | None = None
    term: str | None = None

    # Match "X年级"
    m = re.search(r"((一|二|三|四|五|六)|([1-6]))年级", text)
    if m:
        grade = grade_map[m.group(1)]
    else:
        # Match "X(上/下)" pattern fallback
        m = re.search(r"([一二三四五六])(?=[上下])", text)
        if m:
            grade = grade_map[m.group(1)]

    if re.search(r"上册?|\(上\)", text):
        term = "上学期"
    elif re.search(r"下册?|\(下\)", text):
        term = "下学期"

    return grade, term


def parse_subject(book_name: Any) -> str | None:
    """Identify subject based on predefined keywords."""
    if not isinstance(book_name, str):
        return None

    subjects = [
        "道德与法治",
        "语文",
        "数学",
        "英语",
        "科学",
        "书法",
        "品德与社会",
        "音乐",
        "美术",
        "信息技术",
        "综合实践",
        "体育与健康",
        "体育",
    ]

    for s in subjects:
        if s in book_name:
            # Normalize '体育与健康' to '体育'
            return "体育" if s in ("体育", "体育与健康") else s

    return None


def parse_category(book_name: Any) -> str | None:
    """Categorize as '教辅' or '教材' based on keywords."""
    if not isinstance(book_name, str):
        return None

    keywords = [
        "教参",
        "教师用书",
        "学生活动手册",
        "练习",
        "光盘",
        "学具",
        "字典",
        "教案",
        "辅导",
        "同步",
        "导学",
        "训练",
    ]
    for k in keywords:
        if k in book_name:
            return "教辅"
    return "教材"


def parse_version(book_name: Any, fallback_version: Any) -> Any:
    """Extract publisher version from book name, falling back to provided version if not found."""
    if not isinstance(book_name, str):
        return fallback_version

    text = book_name.replace("（", "(").replace("）", ")")

    explicit_patterns = [
        r"北师大版",
        r"北师大",
        r"人教版",
        r"粤教版",
        r"人音版",
        r"人美版",
        r"语文S版",
        r"语文社S版",
    ]
    for p in explicit_patterns:
        m = re.search(p, text)
        if m:
            v = m.group(0)
            return "北师大版" if v == "北师大" else v

    # Handle "配xxx" pattern (common in supplementary materials)
    m = re.search(r"配\s*([^\)（）]+)", text)
    if m:
        v = m.group(1).strip()
        return "北师大版" if "北师大" in v else v

    for pub in ["湖南美术", "江苏教育", "粤教科技", "粤教育", "粤高教"]:
        if pub in text:
            return pub

    return fallback_version


def main() -> None:
    args = parse_args()
    start_time = time.time()

    # 1. Load and Preprocess Data
    df_src = load_and_normalize_source(args.source)

    # Sort by grade (Custom order: 1-6, others last)
    # Using a temporary column for sorting logic
    df_src["_grade_order"] = df_src["书名"].apply(
        lambda x: {
            "一年级": 1,
            "二年级": 2,
            "三年级": 3,
            "四年级": 4,
            "五年级": 5,
            "六年级": 6,
        }.get(parse_grade_and_term(x)[0] or "", 99)
    )
    df_src = df_src.sort_values(by="_grade_order").drop(columns="_grade_order")

    # 2. Main Output (Template Filling)
    logger.info(f"Processing main output: {args.template} -> {args.output}")

    try:
        wb = load_workbook(args.template)
        ws: Worksheet | None = wb.active
        if ws is None:
            raise ValueError("Template Excel file has no active worksheet.")
    except Exception as e:
        logger.critical(f"Failed to load template: {e}")
        return

    max_col = ws.max_column
    current_row = args.start_row

    # Stats for logging
    stats: dict[str, int] = {"processed": 0, "missing_grade": 0, "missing_subject": 0}

    for idx, (_, r) in enumerate(df_src.iterrows()):
        book = r["书名"]
        grade, term = parse_grade_and_term(book)
        subject = parse_subject(book)
        category = parse_category(book)

        # Logging for data quality issues
        if not grade:
            logger.warning(f"[Row {idx + 1}] Unknown Grade: '{book}'")
            stats["missing_grade"] += 1
        if not subject:
            logger.debug(f"[Row {idx + 1}] Unknown Subject: '{book}'")
            stats["missing_subject"] += 1

        # Extend template styles if necessary
        if current_row > ws.max_row:
            copy_row_style(ws, args.start_row, current_row, max_col)

        # Write Index
        ws.cell(row=current_row, column=1, value=idx + 1)

        values: list[Any] = [
            args.year,
            term,
            grade,
            subject,
            category,
            book,
            r["版别"],
            r["非免费订数"] + r["免费订数"],
            r["单价"],
        ]

        # Write Data (Starting from column 3)
        for col, value in enumerate(values, start=3):
            ws.cell(row=current_row, column=col, value=value)

        current_row += 1
        stats["processed"] += 1

    wb.save(args.output)
    logger.success(f"Main output saved. Stats: {stats}")

    # 3. Side Output (Data Dump)
    logger.info(f"Generating side output: {args.side_output}")

    side_rows: list[dict[str, Any]] = []
    for _, r in df_src.iterrows():
        book = r["书名"]
        grade, _ = parse_grade_and_term(book)

        side_rows.append(
            {
                "版本": parse_version(book, r["版别"]),
                "年级": grade,
                "书名": book,
                "单价": r["单价"],
                "数量": r["非免费订数"] + r["免费订数"],
                "类别": parse_category(book),
                "科目": parse_subject(book),
            }
        )

    pd.DataFrame(side_rows).to_excel(args.side_output, index=False)
    logger.success("Side output saved.")

    logger.info(f"All tasks completed in {time.time() - start_time:.2f} seconds.")
